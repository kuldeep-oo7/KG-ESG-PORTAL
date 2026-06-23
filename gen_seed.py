# -*- coding: utf-8 -*-
"""Generate portal/src/store/SEED.js from the 4 GHG report xlsx files in 'new data/'."""
import openpyxl, os, re, json, io

NEW_DATA = 'new data'
OUT = 'portal/src/store/SEED.js'

FILE_SITE = {
    'GHG-Report-Facets-CY-2023-26.xlsx':         ('KGIPL-03', 'Facets Gems Polishing Works Pvt. Ltd.'),
    'GHG-Report-KG-CY-2023-26.xlsx':             ('KGIPL-02', 'K. Girdharlal International Private Limited'),
    'GHG-Report-Botswana-CY-2023-26.xlsx':       ('KGIPL-05', 'KG Mfg Botswana Proprietary Ltd'),
    'GHG-Report-Mumbai-Office--CY-2023-26.xlsx': ('KGIPL-01', 'Mumbai Office - K. Girdharlal International Pvt'),
}

CAT_MODULE = {
    'Stationary Combustion': ('stationary', 'Scope 1'),
    'Mobile Combustion': ('mobile', 'Scope 1'),
    'Fugitive Emissions': ('fugitive', 'Scope 1'),
    'Purchased Electricity': ('electricity', 'Scope 2'),
    'Heat & Steam': ('heatSteam', 'Scope 2'),
    'Renewable Electricity Generation': ('renewable', 'Scope 2'),
    'Employee Commute': ('employeeCommute', 'Scope 3'),
    'Food Consumption': ('foodConsumption', 'Scope 3'),
    'Purchased Goods': ('purchasedGoods', 'Scope 3'),
    'Transmission & Distribution Loss': ('tdLoss', 'Scope 3'),
    'Upstream Activities': ('upstream', 'Scope 3'),
    'Downstream Activities': ('downstream', 'Scope 3'),
    'Waste Disposal': ('wasteDisposal', 'Scope 3'),
    'Water Supply': ('waterSupply', 'Scope 3'),
    'Water Treatment': ('waterTreatment', 'Scope 3'),
    'Business Travel (Air)': ('businessTravelAir', 'Scope 3'),
    'Business Travel (Sea)': ('businessTravelSea', 'Scope 3'),
    'Business Travel (Land)': ('businessTravelLand', 'Scope 3'),
    'Hotel stay': ('hotelStay', 'Scope 3'),
    'Hotel Stay': ('hotelStay', 'Scope 3'),
}

MONTHS = {m: i for i, m in enumerate(
    ['january','february','march','april','may','june','july','august',
     'september','october','november','december'], 1)}

DATE_RE = re.compile(r'^\s*(\d{1,2})\s*-\s*([A-Za-z]+)\s*-\s*(\d{4})\s*$')

def clean(x):
    if x is None:
        return ''
    return str(x).strip()

def canon_cat(label):
    label = clean(label)
    for key in CAT_MODULE:
        if key.lower() in label.lower():
            return key
    # special: renewable has unicode dash variants
    if 'renewable' in label.lower():
        return 'Renewable Electricity Generation'
    return None

def parse_date(s):
    m = DATE_RE.match(s)
    if not m:
        return None
    d, mon, y = m.group(1), m.group(2).lower(), m.group(3)
    if mon not in MONTHS:
        return None
    return f"{int(y):04d}-{MONTHS[mon]:02d}-{int(d):02d}"

def num(v):
    if v is None or v == '' or v == '-':
        return None
    if isinstance(v, (int, float)):
        return float(v)
    try:
        return float(str(v).replace(',', '').strip())
    except ValueError:
        return None

TYPE_ALIASES = ['Type','type','Vehicle Type','Food Type','Type of Refrigerant','Heat Source',
                'Travel Mode','Mode of Travel','Goods Type','Type of Goods','Name of Country',
                'Country','Standard','Method']
NUM_ALIASES = ['Consumption','consumption','Volume','Volume (m³)','Weight (kg)','Weight (tonnes)',
               'Generation','Nights','Rooms','Tonnes','Distance Travelled','Figure of Distance',
               'Distance (km)','km Travelled']

def descriptor(module, g):
    if module in ('upstream', 'downstream'):
        parts = [g('Type of Vehicle'), g('Type of Fuel'), g('Class'), g('Type')]
    elif module == 'businessTravelAir':
        parts = [g('Mode of Travel'), g('Class'), g('Type')]
    elif module == 'employeeCommute':
        parts = [g('Commute Type'), g('Fuel Type'), g('Vehicle Type')]
    elif module == 'businessTravelLand':
        parts = [g('Type of Fuel'), g('Mode of Travel')]
    elif module == 'hotelStay':
        parts = [g('Name of Country')]
    else:
        parts = []
        for nm in ['Type', 'Select Type', 'Type of Goods', 'Type of Refrigerant',
                   'Type of Fuel', 'Food Type', 'Name of Country', 'Vehicle Type']:
            if g(nm) not in (None, '', '-'):
                parts = [g(nm)]
                break
    parts = [clean(p) for p in parts if clean(p) not in ('', '-')]
    return ' - '.join(parts)

QTY_COLS = ['Consumption','Generation','Kilometers Travelled','Kilo meters Travelled',
            'Distance Travelled','Number of occupied Rooms','Tonnes']

def build_records():
    seed = {}
    cat_totals = {}        # (site, category) -> reported total
    cat_calc = {}          # (site, category) -> summed tco2e
    next_id = 1
    for fname in sorted(os.listdir(NEW_DATA)):
        if not fname.endswith('.xlsx'):
            continue
        site_code, site_name = FILE_SITE[fname]
        seed.setdefault(site_code, {})
        wb = openpyxl.load_workbook(os.path.join(NEW_DATA, fname), data_only=True, read_only=True)
        ws = wb[wb.sheetnames[0]]
        scope = ''
        cat_label = None
        module = None
        hdr = None
        for row in ws.iter_rows(values_only=True):
            c0 = clean(row[0])
            if not c0:
                continue
            up = c0.upper()
            if up.startswith('SCOPE'):
                scope = 'Scope ' + ''.join(ch for ch in c0 if ch.isdigit())
                hdr = None
                continue
            if c0.startswith('Total Emissions'):
                fm = re.search(r'[-+]?\d[\d,]*\.?\d*', c0.split(':', 1)[1] if ':' in c0 else c0)
                t = num(fm.group(0)) if fm else None
                if module is not None and t is not None:
                    cat_totals[(site_code, cat_label)] = t
                hdr = None
                continue
            if c0 == 'Date' and clean(row[1]) == 'Entry Period':
                hdr = [clean(x) for x in row]
                continue
            dt = parse_date(c0)
            if dt is None:
                # category label line
                mk = canon_cat(c0)
                if mk:
                    cat_label = mk
                    module = CAT_MODULE[mk][0]
                hdr = None
                continue
            # ---- data row ----
            if hdr is None or module is None:
                continue
            idx = {h: i for i, h in enumerate(hdr) if h}
            def g(name):
                i = idx.get(name)
                return row[i] if i is not None and i < len(row) else None

            ef = num(g('Emission Factor'))
            ghg = num(g('GHG Emissions (TCO2Eq)'))
            if ghg is None:
                ghg = num(g('Avoided Emissions (TCO2Eq)'))
            qty = None
            for qc in QTY_COLS:
                qv = num(g(qc))
                if qv is not None:
                    qty = qv
                    break
            desc = descriptor(module, g)
            unit = clean(g('Unit of Measurement')) or ('room nights' if module == 'hotelStay' else '')
            source = clean(g('Source')) or 'Defra v 1.0'
            remarks = clean(g('Remarks')) or '-'
            mod_scope = scope or CAT_MODULE.get(cat_label, ('', 'Scope 3'))[1]

            rec = {'id': next_id, 'date': dt}
            next_id += 1
            rec['Entry Period'] = clean(g('Entry Period'))
            rec['period'] = rec['Entry Period']
            rec['site'] = site_code
            rec['siteCode'] = site_code
            rec['site_code'] = site_code
            rec['site_name'] = site_name
            rec['scope'] = 'Scope 2' if module == 'renewable' else mod_scope
            rec['category'] = 'Renewable Electricity Generation' if module == 'renewable' else cat_label
            for k in TYPE_ALIASES:
                rec[k] = desc
            # real country override
            country = clean(g('Name of Country'))
            if country:
                rec['Name of Country'] = country
                rec['Country'] = country
            rec['Source'] = source
            rec['source'] = source
            rec['Unit'] = unit
            rec['unit'] = unit
            rec['Unit of Measurement'] = unit
            qv = qty if qty is not None else 0.0
            for k in NUM_ALIASES:
                rec[k] = qv
            # real physical column overrides
            for col, key in [('Tonnes', 'Tonnes'), ('Distance Travelled', 'Distance Travelled'),
                             ('Generation', 'Generation'),
                             ('Number of occupied Rooms', 'Rooms'),
                             ('Number of Nights Per Room', 'Nights')]:
                rv = num(g(col))
                if rv is not None:
                    rec[key] = rv
            km = num(g('Kilometers Travelled')) or num(g('Kilo meters Travelled'))
            if km is not None:
                rec['Distance (km)'] = km
                rec['km Travelled'] = km
                rec['Distance Travelled'] = km
            rec['Emission Factor'] = ef if ef is not None else 0
            rec['ef'] = rec['Emission Factor']
            rec['tco2e'] = ghg if ghg is not None else 0
            rec['ghg'] = rec['tco2e']
            rec['remarks'] = remarks

            seed[site_code].setdefault(module, []).append(rec)
            key = (site_code, cat_label)
            cat_calc[key] = cat_calc.get(key, 0) + (rec['tco2e'] or 0)
        wb.close()
    return seed, cat_totals, cat_calc

def main():
    seed, cat_totals, cat_calc = build_records()
    # validation
    print('=== PER-CATEGORY VALIDATION (reported vs parsed tco2e) ===')
    bad = 0
    for key in sorted(cat_totals):
        rep = cat_totals[key]
        got = round(cat_calc.get(key, 0), 4)
        ok = abs(rep - got) <= max(0.01, abs(rep) * 0.001)
        if not ok:
            bad += 1
        flag = 'OK ' if ok else '>>>'
        print(f"{flag} {key[0]:9} {key[1]:38} reported={rep:>14.4f}  parsed={got:>14.4f}")
    print(f"\nMismatched categories: {bad}")
    print('\n=== RECORD COUNTS ===')
    total = 0
    for sc in seed:
        cnt = {m: len(seed[sc][m]) for m in seed[sc]}
        n = sum(cnt.values())
        total += n
        print(f"{sc}: {n} records  {cnt}")
    print(f"TOTAL: {total}")
    print('\n=== SITE SCOPE TOTALS (excl renewable) ===')
    for sc in seed:
        t = sum(e['tco2e'] for m in seed[sc] if m != 'renewable' for e in seed[sc][m])
        av = sum(e['tco2e'] for m in seed[sc] if m == 'renewable' for e in seed[sc][m])
        print(f"{sc}: emissions={round(t,3)}  avoided(renewable)={round(av,3)}")

    body = json.dumps(seed, indent=2, ensure_ascii=False)
    out = "// Auto-generated seed data from 'new data/' GHG report workbooks (CY 2023-26)\n"
    out += "export const SEED = " + body + "\n"
    with io.open(OUT, 'w', encoding='utf-8') as f:
        f.write(out)
    print(f"\nWrote {OUT} ({len(out)} bytes)")

if __name__ == '__main__':
    main()
