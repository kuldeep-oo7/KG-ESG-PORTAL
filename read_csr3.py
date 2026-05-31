import openpyxl, warnings, sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
warnings.filterwarnings('ignore')
wb = openpyxl.load_workbook(r'E:\HACKATHON\CSR_Impact_Mapping.xlsx', data_only=True)
for sh in wb.sheetnames:
    ws = wb[sh]
    print(f'=== {sh} ===')
    for row in ws.iter_rows(values_only=True):
        vals = [str(v) if v is not None else '' for v in row]
        if any(v.strip() for v in vals):
            print('\t'.join(vals))
