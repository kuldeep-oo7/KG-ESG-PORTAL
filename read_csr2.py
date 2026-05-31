import openpyxl, warnings, sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
warnings.filterwarnings('ignore')
wb = openpyxl.load_workbook(r'E:\HACKATHON\CSR_Impact_Mapping.xlsx', data_only=True)
print('Sheets:', wb.sheetnames)
for sh in wb.sheetnames:
    ws = wb[sh]
    print(f'\n=== {sh} ===')
    count = 0
    for row in ws.iter_rows(values_only=True):
        vals = [str(v) if v is not None else '' for v in row]
        if any(v.strip() for v in vals):
            print('\t'.join(vals[:20]))
            count += 1
        if count > 200:
            print('...(truncated at 200 rows)')
            break
