import openpyxl
import sys

# Write output to a UTF-8 file to avoid encoding issues
out = open(r"E:\HACKATHON\excel_output.txt", "w", encoding="utf-8")

def p(*args, **kwargs):
    print(*args, **kwargs, file=out)
    try:
        print(*args, **kwargs)  # also to stdout (best effort)
    except UnicodeEncodeError:
        pass  # skip stdout encoding failures

def read_file(filepath, row_limit=None):
    p("=" * 80)
    p(f"FILE: {filepath}")
    p("=" * 80)
    try:
        wb = openpyxl.load_workbook(filepath, data_only=True)
        p(f"Sheets: {wb.sheetnames}")
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            p(f"\n--- Sheet: {sheet_name} (max_row={ws.max_row}, max_col={ws.max_column}) ---")
            row_count = 0
            for row in ws.iter_rows(values_only=True):
                if all(cell is None for cell in row):
                    continue
                line = "\t".join(str(cell) if cell is not None else "" for cell in row)
                p(line)
                row_count += 1
                if row_limit and row_count >= row_limit:
                    p(f"  ... (limited to {row_limit} non-empty rows)")
                    break
        wb.close()
    except Exception as e:
        p(f"ERROR reading {filepath}: {e}")

# File 1
read_file(r"E:\HACKATHON\Book2.xlsx")

# File 2
read_file(r"E:\HACKATHON\esgtech.ai Portal_Scope Wise Data.xlsx")

# File 3
read_file(r"E:\HACKATHON\Food_Consumption_Scope3_Table.xlsx")

# File 4 - large file, limit to 100 rows per sheet
read_file(r"E:\HACKATHON\ghg-conversion-factors-2025.xlsx", row_limit=100)

out.close()
print("\n\nDone. Output written to E:\\HACKATHON\\excel_output.txt")
